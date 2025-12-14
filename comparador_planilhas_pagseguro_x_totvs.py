import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill
import warnings
import sys

# Ignora avisos desnecessários
warnings.filterwarnings("ignore")

class PlanilhaComparatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparador de Planilhas")
        self.root.geometry("800x650")
        
        # Variáveis para armazenar os caminhos dos arquivos
        self.arquivo_operadora = tk.StringVar()
        self.arquivo_totvs = tk.StringVar()
        self.pasta_saida = tk.StringVar()
        self.nome_arquivo = tk.StringVar(value="resultado_final")  # Nome padrão
        
        # Configuração do estilo
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', font=('Arial', 10))
        self.style.configure('TButton', font=('Arial', 10))
        self.style.configure('Title.TLabel', font=('Arial', 16, 'bold'))
        
        # Dicionário para mapear códigos de cliente para bandeiras e tipos
        self.codigo_bandeira_map = {
            '481': {'bandeira': 'VISA', 'tipo': 'debito'},
            '482': {'bandeira': 'VISA', 'tipo': 'credito'},
            '483': {'bandeira': 'MASTERCARD', 'tipo': 'debito'},
            '484': {'bandeira': 'MASTERCARD', 'tipo': 'credito'},
            '485': {'bandeira': 'ELO', 'tipo': 'debito'},
            '486': {'bandeira': 'ELO', 'tipo': 'credito'},
            '487': {'bandeira': 'HIPERCARD', 'tipo': 'credito'},
            '488': {'bandeira': 'AMEX', 'tipo': 'credito'},
            '489': {'bandeira': 'PIX', 'tipo': 'pix'},
            '389': {'bandeira': 'ELO', 'tipo': 'debito'},
            '388': {'bandeira': 'ELO', 'tipo': 'credito'},
            '397': {'bandeira': 'VISA', 'tipo': 'debito'},
            '396': {'bandeira': 'VISA', 'tipo': 'credito'},
            '393': {'bandeira': 'MASTERCARD', 'tipo': 'debito'},
            '394': {'bandeira': 'MASTERCARD', 'tipo': 'credito'},
            '461': {'bandeira': 'PIX', 'tipo': 'pix'}
        }
        
        # Cria a interface
        self.create_widgets()
        
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        title_label = ttk.Label(main_frame, text="Comparador de Planilhas Financeiras Cielo", style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Frame de seleção de arquivos
        file_frame = ttk.LabelFrame(main_frame, text="Seleção de Arquivos", padding="15")
        file_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 20))
        
        # Arquivo da Operadora
        ttk.Label(file_frame, text="Planilha da Operadora:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        operadora_entry = ttk.Entry(file_frame, textvariable=self.arquivo_operadora, width=50)
        operadora_entry.grid(row=0, column=1, sticky="ew")
        ttk.Button(file_frame, text="Selecionar", command=self.select_operadora_file).grid(row=0, column=2, padx=(10, 0))
        
        # Arquivo do Sistema
        ttk.Label(file_frame, text="Planilha do Sistema (TOTVS):").grid(row=1, column=0, sticky="w", padx=(0, 10))
        totvs_entry = ttk.Entry(file_frame, textvariable=self.arquivo_totvs, width=50)
        totvs_entry.grid(row=1, column=1, sticky="ew")
        ttk.Button(file_frame, text="Selecionar", command=self.select_totvs_file).grid(row=1, column=2, padx=(10, 0))
        
        # Pasta de Saída
        ttk.Label(file_frame, text="Pasta para Salvar Resultados:").grid(row=2, column=0, sticky="w", padx=(0, 10))
        output_entry = ttk.Entry(file_frame, textvariable=self.pasta_saida, width=50)
        output_entry.grid(row=2, column=1, sticky="ew")
        ttk.Button(file_frame, text="Selecionar", command=self.select_output_folder).grid(row=2, column=2, padx=(10, 0))
        
        # Nome do Arquivo de Resultado
        ttk.Label(file_frame, text="Nome do Arquivo de Resultado:").grid(row=3, column=0, sticky="w", padx=(0, 10))
        nome_entry = ttk.Entry(file_frame, textvariable=self.nome_arquivo, width=50)
        nome_entry.grid(row=3, column=1, sticky="ew")
        ttk.Label(file_frame, text=".xlsx", font=('Arial', 10, 'bold')).grid(row=3, column=2, sticky="w", padx=(5, 0))
        
        # Frame de configurações
        settings_frame = ttk.LabelFrame(main_frame, text="Configurações", padding="15")
        settings_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 20))
        
        # Opções de processamento
        ttk.Label(settings_frame, text="Tipo de Comparação:").grid(row=0, column=0, sticky="w")
        self.comparison_type = tk.StringVar(value="detalhada")
        ttk.Radiobutton(settings_frame, text="Detalhada (valor a valor)", variable=self.comparison_type, value="detalhada").grid(row=0, column=1, sticky="w")
        ttk.Radiobutton(settings_frame, text="Resumida (por grupo)", variable=self.comparison_type, value="resumida").grid(row=0, column=2, sticky="w")
        
        # Botão de processamento
        process_btn = ttk.Button(main_frame, text="Processar Planilhas", command=self.process_files)
        process_btn.grid(row=3, column=0, columnspan=3, pady=(10, 0))
        
        # Área de logs
        log_frame = ttk.LabelFrame(main_frame, text="Log de Execução", padding="10")
        log_frame.grid(row=4, column=0, columnspan=3, sticky="nsew", pady=(20, 0))
        
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Barra de rolagem
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Configuração de peso para expansão
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)
        file_frame.columnconfigure(1, weight=1)
        
    def select_operadora_file(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha da Operadora",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        if file_path:
            self.arquivo_operadora.set(file_path)
            if not self.pasta_saida.get():
                self.pasta_saida.set(os.path.dirname(file_path))
    
    def select_totvs_file(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha do Sistema (TOTVS)",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        if file_path:
            self.arquivo_totvs.set(file_path)
            if not self.pasta_saida.get():
                self.pasta_saida.set(os.path.dirname(file_path))
    
    def select_output_folder(self):
        folder_path = filedialog.askdirectory(title="Selecione a pasta para salvar os resultados")
        if folder_path:
            self.pasta_saida.set(folder_path)
    
    def log_message(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def validar_nome_arquivo(self, nome):
        """Remove caracteres inválidos do nome do arquivo"""
        caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        nome_limpo = nome
        for char in caracteres_invalidos:
            nome_limpo = nome_limpo.replace(char, '_')
        return nome_limpo.strip()
    
    def process_files(self):
        # Verifica se os arquivos foram selecionados
        if not self.arquivo_operadora.get() or not self.arquivo_totvs.get():
            messagebox.showerror("Erro", "Por favor, selecione ambos os arquivos para comparação.")
            return
        
        if not self.pasta_saida.get():
            messagebox.showerror("Erro", "Por favor, selecione uma pasta para salvar os resultados.")
            return
        
        # Verifica se foi definido um nome para o arquivo
        nome_arquivo = self.nome_arquivo.get().strip()
        if not nome_arquivo:
            messagebox.showerror("Erro", "Por favor, defina um nome para o arquivo de resultado.")
            return
        
        # Limpa o nome do arquivo de caracteres inválidos
        nome_arquivo = self.validar_nome_arquivo(nome_arquivo)
        
        try:
            self.log_message("Iniciando processamento...")
            
            # Processa os arquivos
            df_operadora = self.processar_operadora(self.arquivo_operadora.get())
            df_totvs = self.processar_totvs(self.arquivo_totvs.get())
            
            if df_operadora is None or df_totvs is None:
                self.log_message("\nErro: Não foi possível processar os arquivos. Verifique os logs acima.")
                return
            
            # Gera relatório detalhado
            resultado_detalhado = self.gerar_comparacao_detalhada(df_totvs, df_operadora)
            
            # Formata as datas para o relatório
            resultado_detalhado['Data'] = pd.to_datetime(resultado_detalhado['Data']).dt.strftime('%d/%m/%Y')
            
            # Gera o resumo
            df_resumo = self.gerar_resumo(df_totvs, df_operadora, resultado_detalhado)
            
            # Cria resumo organizado
            df_resumo_organizado = self.criar_resumo_organizado(df_resumo, resultado_detalhado)
            
            # Salva os resultados com o nome personalizado
            resultado_path = os.path.join(self.pasta_saida.get(), f'{nome_arquivo}.xlsx')
            
            with pd.ExcelWriter(resultado_path, engine='openpyxl') as writer:
                df_operadora.to_excel(writer, sheet_name='Operadora Processada', index=False)
                df_totvs.to_excel(writer, sheet_name='TOTVS Processado', index=False)
                resultado_detalhado.to_excel(writer, sheet_name='Comparação Detalhada', index=False)
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                if not df_resumo_organizado.empty:
                    df_resumo_organizado.to_excel(writer, sheet_name='Resumo Filtrável', index=False)
                
                # Aplica formatação condicional
                workbook = writer.book
                worksheet_resumo = writer.sheets['Resumo']
                
                # Determina a coluna do Status (última coluna)
                status_col = len(df_resumo.columns)
                
                # Define os preenchimentos de cor
                verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                # Aplica formatação condicional
                for i, row in enumerate(df_resumo.itertuples(), 2):
                    cell = worksheet_resumo.cell(row=i, column=status_col)
                    if row.Status == "OK":
                        cell.fill = verde
                    elif row.Status == "COM DIFERENÇA":
                        cell.fill = vermelho
            
            self.log_message(f"\nProcessamento concluído com sucesso! Resultados salvos em:\n{resultado_path}")
            
            # Mostra resumo estatístico
            diferenca_a_mais = len([row for _, row in resultado_detalhado.iterrows() if row['A_Mais']])
            diferenca_a_menos = len([row for _, row in resultado_detalhado.iterrows() if row['A_Menos']])
            sem_diferencas = len(df_resumo[df_resumo['Status'] == 'OK'])
            com_diferencas = len(df_resumo[df_resumo['Status'] == 'COM DIFERENÇA'])
            
            self.log_message("\nResumo Estatístico:")
            self.log_message(f"- Total operadora: {len(df_operadora)} transações")
            self.log_message(f"- Total TOTVS: {len(df_totvs)} transações")
            self.log_message(f"- Valores a mais no Sistema: {diferenca_a_mais}")
            self.log_message(f"- Valores a menos no Sistema: {diferenca_a_menos}")
            self.log_message(f"- Combinações sem diferenças: {sem_diferencas}")
            self.log_message(f"- Combinações com diferenças: {com_diferencas}")
            
            messagebox.showinfo("Sucesso", f"Processamento concluído com sucesso!\nArquivo salvo como: {nome_arquivo}.xlsx")
            
        except Exception as e:
            self.log_message(f"\nErro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento:\n{str(e)}")
    
    # Funções de processamento (copiadas do seu código original)
    def formatar_valor(self, valor):
        """Converte valores com vírgula decimal para float"""
        if pd.isna(valor):
            return 0.0
        
        if isinstance(valor, str):
            if '.' in valor and ',' in valor:
                valor = valor.replace('.', '').replace(',', '.')
            elif ',' in valor:
                valor = valor.replace(',', '.')
        return float(valor)

    def normalizar_bandeira(self, nome):
        nome = str(nome).upper()
        if 'MAESTRO' in nome or 'MASTER' in nome or 'MASTERCARD' in nome:
            return 'MASTERCARD'
        elif 'VISA' in nome:
            return 'VISA'
        elif 'ELO' in nome:
            return 'ELO'
        elif 'PIX' in nome:
            return 'PIX'
        return nome.strip()

    def normalizar_tipo(self, tipo):
        tipo = str(tipo).lower()
        if 'débito' in tipo or 'debito' in tipo:
            return 'debito'
        elif 'crédito' in tipo or 'credito' in tipo:
            return 'credito'
        elif 'pix' in tipo:
            return 'pix'
        return tipo.strip()

    def processar_operadora(self, arquivo):
        try:
            self.log_message(f"Processando arquivo da operadora: {arquivo}")
            cols = ['Data da venda', 'Bandeira', 'Forma de pagamento', 'Valor bruto']
            df = pd.read_excel(arquivo, usecols=cols)
            
            df = df.rename(columns={
                'Data da venda': 'Data',
                'Forma de pagamento': 'Tipo',
                'Valor bruto': 'Valor'
            })
            
            df['Data'] = pd.to_datetime(df['Data']).dt.date
            df['Bandeira'] = df['Bandeira'].fillna(df['Tipo'])
            df['Bandeira'] = df['Bandeira'].apply(self.normalizar_bandeira)
            df['Tipo'] = df['Tipo'].apply(self.normalizar_tipo)
            df['Valor'] = df['Valor'].apply(self.formatar_valor)
            
            self.log_message(f"Operadora processada com sucesso. Total de registros: {len(df)}")
            return df[['Data', 'Bandeira', 'Tipo', 'Valor']]
        except Exception as e:
            self.log_message(f"Erro ao processar operadora: {e}")
            return None

    def processar_totvs(self, arquivo):
        try:
            self.log_message(f"Processando arquivo do TOTVS: {arquivo}")
            # Alteração aqui: incluímos a coluna CLIENTE para identificar pelo código
            cols = ['DT. EMISSAO', 'CLIENTE', 'VALOR']
            df = pd.read_excel(arquivo, usecols=cols, header=1)
            
            df['Bandeira'] = ''
            df['Tipo'] = ''
            
            # Nova função que usa o código do cliente para determinar bandeira e tipo
            def extrair_bandeira_tipo_por_codigo(codigo):
                codigo_str = str(codigo).strip()
                if codigo_str in self.codigo_bandeira_map:
                    return (
                        self.codigo_bandeira_map[codigo_str]['bandeira'],
                        self.codigo_bandeira_map[codigo_str]['tipo']
                    )
                else:
                    return 'OUTROS', 'outros'
            
            # Aplica a nova função para cada linha
            for idx, row in df.iterrows():
                bandeira, tipo = extrair_bandeira_tipo_por_codigo(row['CLIENTE'])
                df.at[idx, 'Bandeira'] = bandeira
                df.at[idx, 'Tipo'] = tipo
            
            df['Data'] = pd.to_datetime(df['DT. EMISSAO'], dayfirst=True).dt.date
            df['Valor'] = df['VALOR'].apply(self.formatar_valor)
            
            self.log_message(f"TOTVS processado com sucesso. Total de registros: {len(df)}")
            return df[['Data', 'Bandeira', 'Tipo', 'Valor']]
        except Exception as e:
            self.log_message(f"Erro ao processar TOTVS: {e}")
            return None

    def gerar_comparacao_detalhada(self, df_totvs, df_operadora):
        self.log_message("Gerando comparação detalhada...")
        
        totvs_grouped = df_totvs.groupby(['Data', 'Bandeira', 'Tipo'])['Valor'].apply(list).reset_index()
        operadora_grouped = df_operadora.groupby(['Data', 'Bandeira', 'Tipo'])['Valor'].apply(list).reset_index()
        
        result = []
        
        totvs_keys = set(zip(totvs_grouped['Data'], totvs_grouped['Bandeira'], totvs_grouped['Tipo']))
        operadora_keys = set(zip(operadora_grouped['Data'], operadora_grouped['Bandeira'], operadora_grouped['Tipo']))
        all_keys = totvs_keys.union(operadora_keys)
        
        for data, bandeira, tipo in sorted(all_keys):
            totvs_valores = []
            operadora_valores = []
            
            totvs_row = totvs_grouped[(totvs_grouped['Data'] == data) & 
                                     (totvs_grouped['Bandeira'] == bandeira) & 
                                     (totvs_grouped['Tipo'] == tipo)]
            
            operadora_row = operadora_grouped[(operadora_grouped['Data'] == data) & 
                                             (operadora_grouped['Bandeira'] == bandeira) & 
                                             (operadora_grouped['Tipo'] == tipo)]
            
            if not totvs_row.empty:
                totvs_valores = totvs_row['Valor'].iloc[0]
            
            if not operadora_row.empty:
                operadora_valores = operadora_row['Valor'].iloc[0]
            
            totvs_valores.sort()
            operadora_valores.sort()
            
            valores_a_mais = []
            valores_totvs_restantes = totvs_valores.copy()
            
            for valor_operadora in operadora_valores:
                match_found = False
                for i, valor_totvs in enumerate(valores_totvs_restantes):
                    if valor_totvs == valor_operadora:
                        match_found = True
                        valores_totvs_restantes.pop(i)
                        break
                
                if not match_found:
                    result.append({
                        'Data': data,
                        'Bandeira': bandeira,
                        'Tipo': tipo,
                        'A_Mais': '',
                        'A_Menos': f"{valor_operadora:.2f}".replace('.', ','),
                        'Valor_Sistema': 0,
                        'Valor_Operadora': valor_operadora
                    })
            
            for valor_totvs in valores_totvs_restantes:
                result.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'A_Mais': f"{valor_totvs:.2f}".replace('.', ','),
                    'A_Menos': '',
                    'Valor_Sistema': valor_totvs,
                    'Valor_Operadora': 0
                })
        
        df_result = pd.DataFrame(result)
        if not df_result.empty:
            df_result = df_result.sort_values(['Data', 'Bandeira', 'Tipo'])
        
        self.log_message(f"Comparação detalhada gerada com {len(df_result)} diferenças encontradas.")
        return df_result

    def gerar_resumo(self, df_totvs, df_operadora, resultado_detalhado):
        self.log_message("Gerando resumo da comparação...")
        
        resumo_relatorio = []
        
        def formatar_data(data):
            if isinstance(data, datetime):
                return data.strftime('%d/%m/%Y')
            elif hasattr(data, 'strftime'):
                return data.strftime('%d/%m/%Y')
            else:
                return str(data)
        
        totvs_combos = set(zip([formatar_data(d) for d in df_totvs['Data']], 
                             df_totvs['Bandeira'], 
                             df_totvs['Tipo']))
        operadora_combos = set(zip([formatar_data(d) for d in df_operadora['Data']], 
                                df_operadora['Bandeira'], 
                                df_operadora['Tipo']))
        todas_combos = totvs_combos.union(operadora_combos)
        
        for data, bandeira, tipo in sorted(todas_combos):
            filtro = resultado_detalhado[(resultado_detalhado['Data'] == data) & 
                                        (resultado_detalhado['Bandeira'] == bandeira) & 
                                        (resultado_detalhado['Tipo'] == tipo)]
            
            valores_a_mais = [v for v in filtro['A_Mais'] if v]
            valores_a_menos = [v for v in filtro['A_Menos'] if v]
            
            total_sistema = filtro['Valor_Sistema'].sum() if not filtro.empty else 0
            total_operadora = filtro['Valor_Operadora'].sum() if not filtro.empty else 0
            diferenca_total = total_sistema - total_operadora
            
            resumo_texto = f"dia {data} no {bandeira.lower()} {tipo}\n"
            
            if valores_a_mais:
                resumo_texto += f" a mais : {'/'.join(valores_a_mais)}\n"
            
            if valores_a_menos:
                resumo_texto += f" a menos : {'/'.join(valores_a_menos)}\n"
            
            tem_diferencas = diferenca_total != 0 or valores_a_mais or valores_a_menos
            status = "COM DIFERENÇA" if tem_diferencas else "OK"
            
            resumo_relatorio.append({
                'Data': data,
                'Bandeira': bandeira,
                'Tipo': tipo,
                'Resumo': resumo_texto,
                'Valores_A_Mais': '/'.join(valores_a_mais) if valores_a_mais else '',
                'Valores_A_Menos': '/'.join(valores_a_menos) if valores_a_menos else '',
                'Total_Sistema': total_sistema,
                'Total_Operadora': total_operadora,
                'Diferença_Total': diferenca_total,
                'Status': status
            })
        
        df_resumo = pd.DataFrame(resumo_relatorio)
        self.log_message(f"Resumo gerado com {len(df_resumo)} combinações analisadas.")
        return df_resumo

    def criar_resumo_organizado(self, df_resumo, resultado_detalhado):
        self.log_message("Criando resumo organizado para filtragem...")
        
        linhas = []
        
        for _, row in resultado_detalhado.iterrows():
            data = row['Data']
            bandeira = row['Bandeira']
            tipo = row['Tipo']
            
            if row['A_Mais']:
                linhas.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'Valor': row['A_Mais'],
                    'Tipo_Diferença': 'A_Mais',
                    'Valor_Sistema': row['Valor_Sistema'],
                    'Valor_Operadora': 0
                })
            
            if row['A_Menos']:
                linhas.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'Valor': row['A_Menos'],
                    'Tipo_Diferença': 'A_Menos',
                    'Valor_Sistema': 0,
                    'Valor_Operadora': row['Valor_Operadora']
                })
        
        df_organizado = pd.DataFrame(linhas)
        
        if not df_organizado.empty:
            df_organizado = df_organizado.sort_values(['Data', 'Bandeira', 'Tipo', 'Tipo_Diferença'])
            df_organizado['Valor_Numérico'] = df_organizado['Valor'].str.replace(',', '.').astype(float)
        
        self.log_message(f"Resumo organizado criado com {len(df_organizado)} diferenças listadas.")
        return df_organizado

if __name__ == "__main__":
    root = tk.Tk()
    app = PlanilhaComparatorApp(root)
    root.mainloop()