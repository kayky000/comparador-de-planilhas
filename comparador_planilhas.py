import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill
import warnings

warnings.filterwarnings("ignore")

# Configurações de aparência
ctk.set_appearance_mode("dark")  # Modes: "System", "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"

class PlanilhaComparatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparador de Planilhas Financeiras")
        self.root.geometry("1000x800")
        
        # Variáveis
        self.arquivo_operadora = ctk.StringVar()
        self.arquivo_totvs = ctk.StringVar()
        self.pasta_saida = ctk.StringVar()
        self.nome_arquivo = ctk.StringVar(value="resultado_final")
        
        # Dicionário de códigos
        self.codigo_bandeira_map = {
            '481': {'bandeira': 'PAGSEGURO VISA', 'tipo': 'debito'},
            '482': {'bandeira': 'PAGSEGURO VISA', 'tipo': 'credito'},
            '483': {'bandeira': 'PAGSEGURO MASTERCARD', 'tipo': 'debito'},
            '484': {'bandeira': 'PAGSEGURO MASTERCARD', 'tipo': 'credito'},
            '485': {'bandeira': 'PAGSEGURO ELO', 'tipo': 'debito'},
            '486': {'bandeira': 'PAGSEGURO ELO', 'tipo': 'credito'},
            '487': {'bandeira': 'PAGSEGURO HIPERCARD', 'tipo': 'credito'},
            '488': {'bandeira': 'PAGSEGURO AMEX', 'tipo': 'credito'},
            '489': {'bandeira': 'PAGSEGURO PIX', 'tipo': 'pix'},
            '389': {'bandeira': 'ELO', 'tipo': 'debito'},
            '388': {'bandeira': 'ELO', 'tipo': 'credito'},
            '397': {'bandeira': 'VISA', 'tipo': 'debito'},
            '396': {'bandeira': 'VISA', 'tipo': 'credito'},
            '393': {'bandeira': 'MASTERCARD', 'tipo': 'debito'},
            '394': {'bandeira': 'MASTERCARD', 'tipo': 'credito'},
            '461': {'bandeira': 'PIX', 'tipo': 'pix'}
        }
        
        self.create_widgets()
        
    def create_widgets(self):
        # Container principal com padding
        main_container = ctk.CTkFrame(self.root, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        header_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            header_frame, 
            text="📊 Comparador de Planilhas Financeiras Cielo",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack()
        
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Compare dados da Operadora com o Sistema TOTVS de forma automatizada",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        subtitle_label.pack(pady=(5, 0))
        
        # Frame de arquivos
        files_frame = ctk.CTkFrame(main_container)
        files_frame.pack(fill="x", pady=(0, 10))
        
        files_title = ctk.CTkLabel(
            files_frame,
            text="Seleção de Arquivos",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        files_title.grid(row=0, column=0, columnspan=3, sticky="w", padx=20, pady=(15, 10))
        
        # Operadora
        self.create_file_row(
            files_frame, 
            row=1,
            label="📄 Planilha da Operadora:",
            variable=self.arquivo_operadora,
            command=self.select_operadora_file
        )
        
        # TOTVS
        self.create_file_row(
            files_frame,
            row=2,
            label="💼 Planilha do Sistema (TOTVS):",
            variable=self.arquivo_totvs,
            command=self.select_totvs_file
        )
        
        # Pasta de saída
        self.create_file_row(
            files_frame,
            row=3,
            label="📁 Pasta para Salvar Resultados:",
            variable=self.pasta_saida,
            command=self.select_output_folder,
            is_folder=True
        )
        
        # Nome do arquivo
        ctk.CTkLabel(
            files_frame,
            text="📝 Nome do Arquivo:",
            font=ctk.CTkFont(size=13)
        ).grid(row=4, column=0, sticky="w", padx=20, pady=10)
        
        nome_frame = ctk.CTkFrame(files_frame, fg_color="transparent")
        nome_frame.grid(row=4, column=1, sticky="ew", padx=10, pady=10)
        
        self.nome_entry = ctk.CTkEntry(
            nome_frame,
            textvariable=self.nome_arquivo,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.nome_entry.pack(side="left", fill="x", expand=True)
        
        ctk.CTkLabel(
            nome_frame,
            text=".xlsx",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="left", padx=(5, 0))
        
        files_frame.columnconfigure(1, weight=1)
        
        # Frame de configurações
        config_frame = ctk.CTkFrame(main_container)
        config_frame.pack(fill="x", pady=(0, 8))
        
        config_title = ctk.CTkLabel(
            config_frame,
            text="⚙️ Configurações",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        config_title.grid(row=0, column=0, columnspan=3, sticky="w", padx=20, pady=(15, 10))
        
        ctk.CTkLabel(
            config_frame,
            text="Tipo de Comparação:",
            font=ctk.CTkFont(size=13)
        ).grid(row=1, column=0, sticky="w", padx=20, pady=10)
        
        self.comparison_type = ctk.StringVar(value="detalhada")
        
        radio_frame = ctk.CTkFrame(config_frame, fg_color="transparent")
        radio_frame.grid(row=1, column=1, sticky="w", padx=10, pady=10)
        
        ctk.CTkRadioButton(
            radio_frame,
            text="🔍 Detalhada (valor a valor)",
            variable=self.comparison_type,
            value="detalhada",
            font=ctk.CTkFont(size=12)
        ).pack(side="left", padx=(0, 20))
        
        ctk.CTkRadioButton(
            radio_frame,
            text="📑 Resumida (por grupo)",
            variable=self.comparison_type,
            value="resumida",
            font=ctk.CTkFont(size=12)
        ).pack(side="left")
        
        # Botão de processar
        self.process_button = ctk.CTkButton(
            main_container,
            text="🚀 Processar Planilhas",
            command=self.process_files,
            height=38,
            font=ctk.CTkFont(size=13, weight="bold"),
            corner_radius=8
        )
        self.process_button.pack(fill="x", pady=(0, 8))
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(main_container)
        self.progress_bar.pack(fill="x", pady=(0, 8))
        self.progress_bar.set(0)
        
        # Frame de logs
        log_frame = ctk.CTkFrame(main_container)
        log_frame.pack(fill="both", expand=True)
        
        log_title = ctk.CTkLabel(
            log_frame,
            text="📋 Log de Execução",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        log_title.pack(anchor="w", padx=20, pady=(15, 10))
        
        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(family="Consolas", size=11),
            wrap="word"
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Botão de tema
        self.theme_button = ctk.CTkButton(
            main_container,
            text="🌓 Alternar Tema",
            command=self.toggle_theme,
            width=130,
            height=30,
            font=ctk.CTkFont(size=12),
            corner_radius=8
        )
        self.theme_button.pack(pady=(5, 0))
        
    def create_file_row(self, parent, row, label, variable, command, is_folder=False):
        ctk.CTkLabel(
            parent,
            text=label,
            font=ctk.CTkFont(size=13)
        ).grid(row=row, column=0, sticky="w", padx=20, pady=10)
        
        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            height=32,
            font=ctk.CTkFont(size=11)
        )
        entry.grid(row=row, column=1, sticky="ew", padx=10, pady=8)
        
        icon = "📁" if is_folder else "📂"
        button = ctk.CTkButton(
            parent,
            text=f"{icon} Selecionar",
            command=command,
            width=110,
            height=32,
            font=ctk.CTkFont(size=11)
        )
        button.grid(row=row, column=2, padx=20, pady=10)
    
    def toggle_theme(self):
        current = ctk.get_appearance_mode()
        new_mode = "Light" if current == "Dark" else "Dark"
        ctk.set_appearance_mode(new_mode)
        self.log_message(f"Tema alterado para: {new_mode}")
    
    def select_operadora_file(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha da Operadora",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        if file_path:
            self.arquivo_operadora.set(file_path)
            if not self.pasta_saida.get():
                self.pasta_saida.set(os.path.dirname(file_path))
            self.log_message(f"✅ Arquivo da Operadora selecionado")
    
    def select_totvs_file(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha do Sistema (TOTVS)",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        if file_path:
            self.arquivo_totvs.set(file_path)
            if not self.pasta_saida.get():
                self.pasta_saida.set(os.path.dirname(file_path))
            self.log_message(f"✅ Arquivo TOTVS selecionado")
    
    def select_output_folder(self):
        folder_path = filedialog.askdirectory(title="Selecione a pasta para salvar os resultados")
        if folder_path:
            self.pasta_saida.set(folder_path)
            self.log_message(f"✅ Pasta de saída definida")
    
    def log_message(self, message):
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()
    
    def validar_nome_arquivo(self, nome):
        caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        nome_limpo = nome
        for char in caracteres_invalidos:
            nome_limpo = nome_limpo.replace(char, '_')
        return nome_limpo.strip()
    
    def process_files(self):
        if not self.arquivo_operadora.get() or not self.arquivo_totvs.get():
            messagebox.showerror("Erro", "Por favor, selecione ambos os arquivos para comparação.")
            return
        
        if not self.pasta_saida.get():
            messagebox.showerror("Erro", "Por favor, selecione uma pasta para salvar os resultados.")
            return
        
        nome_arquivo = self.nome_arquivo.get().strip()
        if not nome_arquivo:
            messagebox.showerror("Erro", "Por favor, defina um nome para o arquivo de resultado.")
            return
        
        nome_arquivo = self.validar_nome_arquivo(nome_arquivo)
        
        try:
            self.process_button.configure(state="disabled", text="⏳ Processando...")
            self.progress_bar.set(0.1)
            self.log_message("\n" + "="*60)
            self.log_message("🚀 Iniciando processamento...")
            self.log_message("="*60)
            
            # Processa os arquivos
            df_operadora = self.processar_operadora(self.arquivo_operadora.get())
            self.progress_bar.set(0.3)
            
            df_totvs = self.processar_totvs(self.arquivo_totvs.get())
            self.progress_bar.set(0.5)
            
            if df_operadora is None or df_totvs is None:
                self.log_message("\n❌ Erro: Não foi possível processar os arquivos.")
                return
            
            # Gera relatório detalhado
            resultado_detalhado = self.gerar_comparacao_detalhada(df_totvs, df_operadora)
            self.progress_bar.set(0.7)
            
            resultado_detalhado['Data'] = pd.to_datetime(resultado_detalhado['Data']).dt.strftime('%d/%m/%Y')
            
            # Gera o resumo
            df_resumo = self.gerar_resumo(df_totvs, df_operadora, resultado_detalhado)
            df_resumo_organizado = self.criar_resumo_organizado(df_resumo, resultado_detalhado)
            self.progress_bar.set(0.85)
            
            # Salva os resultados
            resultado_path = os.path.join(self.pasta_saida.get(), f'{nome_arquivo}.xlsx')
            
            with pd.ExcelWriter(resultado_path, engine='openpyxl') as writer:
                df_operadora.to_excel(writer, sheet_name='Operadora Processada', index=False)
                df_totvs.to_excel(writer, sheet_name='TOTVS Processado', index=False)
                resultado_detalhado.to_excel(writer, sheet_name='Comparação Detalhada', index=False)
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                
                if not df_resumo_organizado.empty:
                    df_resumo_organizado.to_excel(writer, sheet_name='Resumo Filtrável', index=False)
                
                # Formatação condicional
                workbook = writer.book
                worksheet_resumo = writer.sheets['Resumo']
                status_col = len(df_resumo.columns)
                
                verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                for i, row in enumerate(df_resumo.itertuples(), 2):
                    cell = worksheet_resumo.cell(row=i, column=status_col)
                    if row.Status == "OK":
                        cell.fill = verde
                    elif row.Status == "COM DIFERENÇA":
                        cell.fill = vermelho
            
            self.progress_bar.set(1.0)
            
            # Estatísticas
            diferenca_a_mais = len([row for _, row in resultado_detalhado.iterrows() if row['A_Mais_Sistema']])
            diferenca_a_menos = len([row for _, row in resultado_detalhado.iterrows() if row['A_Menos_Sistema']])
            sem_diferencas = len(df_resumo[df_resumo['Status'] == 'OK'])
            com_diferencas = len(df_resumo[df_resumo['Status'] == 'COM DIFERENÇA'])
            
            self.log_message("\n" + "="*60)
            self.log_message("📊 RESUMO ESTATÍSTICO")
            self.log_message("="*60)
            self.log_message(f"📄 Total de transações na Operadora: {len(df_operadora)}")
            self.log_message(f"💼 Total de transações no Sistema: {len(df_totvs)}")
            self.log_message("-"*60)
            self.log_message(f"⬆️  Valores A MAIS no Sistema: {diferenca_a_mais}")
            self.log_message(f"    (lançados no Sistema mas não encontrados na Operadora)")
            self.log_message(f"⬇️  Valores A MENOS no Sistema: {diferenca_a_menos}")
            self.log_message(f"    (existem na Operadora mas não foram lançados no Sistema)")
            self.log_message("-"*60)
            self.log_message(f"✅ Combinações sem diferenças: {sem_diferencas}")
            self.log_message(f"⚠️  Combinações com diferenças: {com_diferencas}")
            self.log_message("="*60)
            self.log_message(f"\n💾 Arquivo salvo em:\n{resultado_path}\n")
            
            messagebox.showinfo("✅ Sucesso", f"Processamento concluído com sucesso!\n\nArquivo salvo como:\n{nome_arquivo}.xlsx")
            
        except Exception as e:
            self.log_message(f"\n❌ Erro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento:\n{str(e)}")
        finally:
            self.process_button.configure(state="normal", text="🚀 Processar Planilhas")
    
    # Funções de processamento (mantidas do código original)
    def formatar_valor(self, valor):
        if pd.isna(valor):
            return 0.0
        if isinstance(valor, str):
            if '.' in valor and ',' in valor:
                valor = valor.replace('.', '').replace(',', '.')
            elif ',' in valor:
                valor = valor.replace(',', '.')
        return float(valor)

    def normalizar_bandeira(self, nome):
        nome = str(nome).upper()
        if 'MAESTRO' in nome or 'MASTER' in nome or 'MASTERCARD' in nome:
            return 'MASTERCARD'
        elif 'VISA' in nome:
            return 'VISA'
        elif 'ELO' in nome:
            return 'ELO'
        elif 'PIX' in nome:
            return 'PIX'
        return nome.strip()

    def normalizar_tipo(self, tipo):
        tipo = str(tipo).lower()
        if 'débito' in tipo or 'debito' in tipo:
            return 'debito'
        elif 'crédito' in tipo or 'credito' in tipo:
            return 'credito'
        elif 'pix' in tipo:
            return 'pix'
        return tipo.strip()

    def processar_operadora(self, arquivo):
        try:
            self.log_message(f"📂 Processando arquivo da operadora...")
            cols = ['Data da venda', 'Bandeira', 'Forma de pagamento', 'Valor bruto']
            df = pd.read_excel(arquivo, usecols=cols)
            
            df = df.rename(columns={
                'Data da venda': 'Data',
                'Forma de pagamento': 'Tipo',
                'Valor bruto': 'Valor'
            })
            
            df['Data'] = pd.to_datetime(df['Data'], dayfirst=True).dt.date
            df['Bandeira'] = df['Bandeira'].fillna(df['Tipo'])
            df['Bandeira'] = df['Bandeira'].apply(self.normalizar_bandeira)
            df['Tipo'] = df['Tipo'].apply(self.normalizar_tipo)
            df['Valor'] = df['Valor'].apply(self.formatar_valor)
            
            self.log_message(f"✅ Operadora processada: {len(df)} registros")
            return df[['Data', 'Bandeira', 'Tipo', 'Valor']]
        except Exception as e:
            self.log_message(f"❌ Erro ao processar operadora: {e}")
            return None

    def processar_totvs(self, arquivo):
        try:
            self.log_message(f"📂 Processando arquivo do TOTVS...")
            cols = ['DT. EMISSAO', 'CLIENTE', 'VALOR']
            df = pd.read_excel(arquivo, usecols=cols, header=1)
            
            df['Bandeira'] = ''
            df['Tipo'] = ''
            
            def extrair_bandeira_tipo_por_codigo(codigo):
                codigo_str = str(codigo).strip()
                if codigo_str in self.codigo_bandeira_map:
                    return (
                        self.codigo_bandeira_map[codigo_str]['bandeira'],
                        self.codigo_bandeira_map[codigo_str]['tipo']
                    )
                else:
                    return 'OUTROS', 'outros'
            
            for idx, row in df.iterrows():
                bandeira, tipo = extrair_bandeira_tipo_por_codigo(row['CLIENTE'])
                df.at[idx, 'Bandeira'] = bandeira
                df.at[idx, 'Tipo'] = tipo
            
            df['Data'] = pd.to_datetime(df['DT. EMISSAO'], dayfirst=True).dt.date
            df['Valor'] = df['VALOR'].apply(self.formatar_valor)
            
            self.log_message(f"✅ TOTVS processado: {len(df)} registros")
            return df[['Data', 'Bandeira', 'Tipo', 'Valor']]
        except Exception as e:
            self.log_message(f"❌ Erro ao processar TOTVS: {e}")
            return None

    def gerar_comparacao_detalhada(self, df_totvs, df_operadora):
        self.log_message("🔍 Gerando comparação detalhada...")
        
        totvs_grouped = df_totvs.groupby(['Data', 'Bandeira', 'Tipo'])['Valor'].apply(list).reset_index()
        operadora_grouped = df_operadora.groupby(['Data', 'Bandeira', 'Tipo'])['Valor'].apply(list).reset_index()
        
        result = []
        
        totvs_keys = set(zip(totvs_grouped['Data'], totvs_grouped['Bandeira'], totvs_grouped['Tipo']))
        operadora_keys = set(zip(operadora_grouped['Data'], operadora_grouped['Bandeira'], operadora_grouped['Tipo']))
        all_keys = totvs_keys.union(operadora_keys)
        
        for data, bandeira, tipo in sorted(all_keys):
            totvs_valores = []
            operadora_valores = []
            
            totvs_row = totvs_grouped[(totvs_grouped['Data'] == data) & 
                                     (totvs_grouped['Bandeira'] == bandeira) & 
                                     (totvs_grouped['Tipo'] == tipo)]
            
            operadora_row = operadora_grouped[(operadora_grouped['Data'] == data) & 
                                             (operadora_grouped['Bandeira'] == bandeira) & 
                                             (operadora_grouped['Tipo'] == tipo)]
            
            if not totvs_row.empty:
                totvs_valores = totvs_row['Valor'].iloc[0].copy()
            
            if not operadora_row.empty:
                operadora_valores = operadora_row['Valor'].iloc[0].copy()
            
            valores_sistema_restantes = totvs_valores.copy()
            valores_operadora_restantes = operadora_valores.copy()
            
            for valor_operadora in operadora_valores[:]:
                for i, valor_sistema in enumerate(valores_sistema_restantes):
                    if valor_sistema == valor_operadora:
                        valores_sistema_restantes.pop(i)
                        if valor_operadora in valores_operadora_restantes:
                            valores_operadora_restantes.remove(valor_operadora)
                        break
            
            for valor_sistema in valores_sistema_restantes:
                result.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'A_Mais_Sistema': f"{valor_sistema:.2f}".replace('.', ','),
                    'A_Menos_Sistema': '',
                    'Valor_Sistema': valor_sistema,
                    'Valor_Operadora': 0,
                    'Observação': 'Valor lançado no Sistema mas não encontrado na Operadora'
                })
            
            for valor_operadora in valores_operadora_restantes:
                result.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'A_Mais_Sistema': '',
                    'A_Menos_Sistema': f"{valor_operadora:.2f}".replace('.', ','),
                    'Valor_Sistema': 0,
                    'Valor_Operadora': valor_operadora,
                    'Observação': 'Valor na Operadora mas não lançado no Sistema'
                })
        
        df_result = pd.DataFrame(result)
        if not df_result.empty:
            df_result = df_result.sort_values(['Data', 'Bandeira', 'Tipo'])
        
        self.log_message(f"✅ Comparação concluída: {len(df_result)} diferenças")
        return df_result

    def gerar_resumo(self, df_totvs, df_operadora, resultado_detalhado):
        self.log_message("📊 Gerando resumo...")
        
        resumo_relatorio = []
        
        def formatar_data(data):
            if isinstance(data, datetime):
                return data.strftime('%d/%m/%Y')
            elif hasattr(data, 'strftime'):
                return data.strftime('%d/%m/%Y')
            else:
                return str(data)
        
        totvs_combos = set(zip([formatar_data(d) for d in df_totvs['Data']], 
                             df_totvs['Bandeira'], 
                             df_totvs['Tipo']))
        operadora_combos = set(zip([formatar_data(d) for d in df_operadora['Data']], 
                                df_operadora['Bandeira'], 
                                df_operadora['Tipo']))
        todas_combos = totvs_combos.union(operadora_combos)
        
        for data, bandeira, tipo in sorted(todas_combos):
            filtro = resultado_detalhado[(resultado_detalhado['Data'] == data) & 
                                        (resultado_detalhado['Bandeira'] == bandeira) & 
                                        (resultado_detalhado['Tipo'] == tipo)]
            
            valores_a_mais = [v for v in filtro['A_Mais_Sistema'] if v]
            valores_a_menos = [v for v in filtro['A_Menos_Sistema'] if v]
            
            total_sistema = filtro['Valor_Sistema'].sum() if not filtro.empty else 0
            total_operadora = filtro['Valor_Operadora'].sum() if not filtro.empty else 0
            diferenca_total = total_sistema - total_operadora
            
            resumo_texto = f"dia {data} no {bandeira.lower()} {tipo}\n"
            
            if valores_a_mais:
                resumo_texto += f" A MAIS no Sistema: {'/'.join(valores_a_mais)}\n"
            
            if valores_a_menos:
                resumo_texto += f" A MENOS no Sistema (falta lançar): {'/'.join(valores_a_menos)}\n"
            
            tem_diferencas = diferenca_total != 0 or valores_a_mais or valores_a_menos
            status = "COM DIFERENÇA" if tem_diferencas else "OK"
            
            resumo_relatorio.append({
                'Data': data,
                'Bandeira': bandeira,
                'Tipo': tipo,
                'Resumo': resumo_texto,
                'Valores_A_Mais_Sistema': '/'.join(valores_a_mais) if valores_a_mais else '',
                'Valores_A_Menos_Sistema': '/'.join(valores_a_menos) if valores_a_menos else '',
                'Total_Sistema': total_sistema,
                'Total_Operadora': total_operadora,
                'Diferença_Total': diferenca_total,
                'Status': status
            })
        
        df_resumo = pd.DataFrame(resumo_relatorio)
        self.log_message(f"✅ Resumo gerado: {len(df_resumo)} combinações")
        return df_resumo

    def criar_resumo_organizado(self, df_resumo, resultado_detalhado):
        self.log_message("📋 Criando resumo organizado...")
        
        linhas = []
        
        for _, row in resultado_detalhado.iterrows():
            data = row['Data']
            bandeira = row['Bandeira']
            tipo = row['Tipo']
            
            if row['A_Mais_Sistema']:
                linhas.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'Valor': row['A_Mais_Sistema'],
                    'Tipo_Diferença': 'A_Mais',
                    'Valor_Sistema': row['Valor_Sistema'],
                    'Valor_Operadora': 0,
                    'Descrição': 'Lançado no Sistema, não encontrado na Operadora'
                })
            
            if row['A_Menos_Sistema']:
                linhas.append({
                    'Data': data,
                    'Bandeira': bandeira,
                    'Tipo': tipo,
                    'Valor': row['A_Menos_Sistema'],
                    'Tipo_Diferença': 'A_Menos',
                    'Valor_Sistema': 0,
                    'Valor_Operadora': row['Valor_Operadora'],
                    'Descrição': 'Encontrado na Operadora, não lançado no Sistema'
                })
        
        df_organizado = pd.DataFrame(linhas)
        
        if not df_organizado.empty:
            df_organizado = df_organizado.sort_values(['Data', 'Bandeira', 'Tipo', 'Tipo_Diferença'])
            df_organizado['Valor_Numérico'] = df_organizado['Valor'].str.replace(',', '.').astype(float)
        
        self.log_message(f"✅ Resumo organizado: {len(df_organizado)} diferenças listadas")
        return df_organizado

if __name__ == "__main__":
    root = ctk.CTk()
    app = PlanilhaComparatorApp(root)
    root.mainloop()